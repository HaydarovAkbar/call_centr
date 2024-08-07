from rest_framework.generics import ListAPIView, CreateAPIView, RetrieveAPIView, UpdateAPIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser, FileUploadParser
from rest_framework.response import Response
import openpyxl as px
from django.http import FileResponse
from django.db.models import Count
import os
from rest_framework.permissions import IsAuthenticated
from django.conf import settings
from django.contrib.auth.models import User
import datetime

from .serializers import AppealSerializer, AppealCreateSerializer, FAQSerializer, ChangeAppealStatusSerializers, \
    StatusSerializer, AnswersSerializer
from .models import Appeal, FAQ, Status, Answers
from .permissions import IsCallCenter
from .pagination import TenPagination
from .filters import AppealDatetimeFilters


class AppealListView(ListAPIView):
    queryset = Appeal.objects.all().order_by('-id')
    serializer_class = AppealSerializer
    filter_backends = [DjangoFilterBackend, AppealDatetimeFilters]
    filterset_fields = ['user', 'phone_number', 'district', 'region', 'status']
    permission_classes = [IsAuthenticated]
    http_method_names = ['get', ]
    pagination_class = TenPagination


class AppealCreateView(CreateAPIView):
    queryset = Appeal.objects.all()
    serializer_class = AppealCreateSerializer
    permission_classes = [IsCallCenter, IsAuthenticated]
    http_method_names = ['post', ]

    parser_classes = [MultiPartParser, FormParser, FileUploadParser]  # JSONParser

    def create(self, request, *args, **kwargs):
        user = request.user
        request.data['user'] = user.id
        return super().create(request, *args, **kwargs)


class AppealRetrieveView(RetrieveAPIView):
    queryset = Appeal.objects.all()
    serializer_class = AppealSerializer
    # permission_classes = [IsAuthenticated, IsCallCenter]
    http_method_names = ['get', ]


class AppealUpdateView(UpdateAPIView):
    queryset = Appeal.objects.all()
    serializer_class = AppealCreateSerializer
    # permission_classes = [IsAuthenticated, IsCallCenter]
    http_method_names = ['patch', ]


class FAQListView(ListAPIView):
    queryset = FAQ.objects.all().order_by('order')
    serializer_class = FAQSerializer
    http_method_names = ['get', ]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['title', 'question']


class ImportAppealView(ListAPIView):
    queryset = Appeal.objects.all()
    serializer_class = AppealCreateSerializer
    # permission_classes = [IsCallCenter, IsAuthenticated]
    http_method_names = ['get', ]
    filter_backends = [DjangoFilterBackend, AppealDatetimeFilters]
    filterset_fields = ['user', 'phone_number', 'district', 'region']

    def list(self, request, *args, **kwargs):
        filter_queryset = self.filter_queryset(self.get_queryset()).order_by('-id')
        excel_name = 'report.xlsx'
        excel_path = os.path.join(settings.STATIC_ROOT, excel_name)
        book = px.load_workbook(excel_path)
        wb = book.get_sheet_by_name('book')
        # wb.cell(3, 10,
        #         f'{now().date().year} yil {now().date().day} may hour {now().hour}:{now().strftime("%M")} holatiga')
        row, count = 4, 1
        for query in filter_queryset:
            wb.cell(row, 2, str(count))
            wb.cell(row, 3, query.region.title if query.region else None)
            wb.cell(row, 4, query.district.title if query.district else None)
            wb.cell(row, 5, query.app_name)
            wb.cell(row, 6, query.phone_number)
            wb.cell(row, 7, (query.user.first_name + ' ' + query.user.last_name) if query.user else None)
            wb.cell(row, 8, query.app_datetime.strftime('%Y-%m-%d %H:%M:%S'))
            wb.cell(row, 9, query.get_voice_url() if query.voice else None)
            row += 1
            count += 1
        book.save(f"static\data.xlsx")
        return FileResponse(open('static\data.xlsx', 'rb'))


class StatsAppealView(ListAPIView):
    queryset = Appeal.objects.all()
    serializer_class = AppealSerializer
    filter_backends = [DjangoFilterBackend, AppealDatetimeFilters]
    filterset_fields = ['user', 'phone_number', 'district', 'region', 'faq']
    http_method_names = ['get', ]

    def list(self, request, *args, **kwargs):
        filter_queryset = self.filter_queryset(self.get_queryset())
        data = {'count': filter_queryset.count()}

        # top 5 appeals by region and faq
        data['top_region'] = filter_queryset.values('region__title').annotate(
            count=Count('region')).order_by('-count')[:5]
        data['top_faq'] = filter_queryset.values('faq__title').annotate(
            count=Count('faq')).order_by('-count')[:5]

        row, table = list(), list()
        for date in range(7):
            date = datetime.datetime.now() - datetime.timedelta(days=date)
            row.append(date.strftime('%Y-%m-%d'))
            table.append(filter_queryset.filter(app_datetime__date=date).count())
        data['daily'] = {
            'row': row[::-1],
            'data': table[::-1]
        }

        # top 5 appeals by user
        data['top_user'] = filter_queryset.values('user__first_name', 'user__last_name').annotate(
            count=Count('user')).order_by('-count')[:5]
        # top 5 appeals by status
        data['top_status'] = filter_queryset.values('status__title').annotate(
            count=Count('status')).order_by('-count')[:5]
        data['top_answers'] = filter_queryset.values('answers__title').annotate(
            count=Count('answers')).order_by('-count')

        # How many times each user has talked should be displayed in the status section
        charts = {}
        # filter user if group in call center
        users = User.objects.filter(groups__name='call_center')
        for user in users:
            user_fullname = user.first_name + ' ' + user.last_name
            # filter queryset by status and user
            user_queryset = filter_queryset.filter(user=user)
            d = []
            for status in Status.objects.all():
                d.append({
                    'status': status.title,
                    'count': user_queryset.filter(status=status).count()
                })
            charts[user_fullname] = d
        data['charts'] = charts
        return Response(data)


class StatusListView(ListAPIView):
    queryset = Status.objects.all().order_by('order')
    serializer_class = StatusSerializer
    http_method_names = ['get', ]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['title', ]
    permission_classes = [IsAuthenticated, ]


class ChangeAppealStatusView(UpdateAPIView):
    queryset = Appeal.objects.all()
    serializer_class = ChangeAppealStatusSerializers
    permission_classes = [IsAuthenticated, IsCallCenter]
    http_method_names = ['patch', 'put', ]


class AnswersListView(ListAPIView):
    queryset = Answers.objects.all().order_by('order')
    serializer_class = AnswersSerializer
    http_method_names = ['get', ]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['title', 'question']
    permission_classes = [IsAuthenticated, ]